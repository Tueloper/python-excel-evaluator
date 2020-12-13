import openpyxl as xl
from openpyxl.chart import  BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['SalesOrders']

    for row in range(2, sheet.max_row + 1):
        units = sheet.cell(row, 5)
        unit_cost = sheet.cell(row, 6)
        total_price = units.value * unit_cost.value
        total_price_cell = sheet.cell(row, 7)
        total_price_cell.value = total_price

    # creating a chart
    cell4 = Reference(sheet,
                      min_row=2,
                      max_row=sheet.max_row,
                      min_col=7,
                      max_col=7)

    chart = BarChart()
    chart.add_data(cell4)

    sheet.add_chart(chart, 'a50')

    # save
    wb.save(filename)


process_workbook('SampleData.xlsx')
