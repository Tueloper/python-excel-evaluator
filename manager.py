import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import json


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['staff']

    cells = []

    for row in range(2, sheet.max_row + 1):
        cell = {
            'managerNumber': sheet.cell(row, 13).value,
            'name': sheet.cell(row, 14).value,
            'email': sheet.cell(row, 15).value,
            'peopleManagerEmail': sheet.cell(row, 13).value,
            'unit': sheet.cell(row, 11).value
        }
        cells.append(cell)

    with open('manager.json', 'w') as outfile:
        json.dump(cells, outfile)
    # print(cells)

    # save
    # wb.save(filename)


# print(cells)

process_workbook('staff.xlsx')
