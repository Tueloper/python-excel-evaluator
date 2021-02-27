import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import json


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['staff']

    cells = []

    for row in range(2, sheet.max_row + 1):
        cell = {
            'employeeReference': sheet.cell(row, 1).value,
            'firstName': sheet.cell(row, 5).value,
            'lastName': sheet.cell(row, 6).value,
            'nationality': sheet.cell(row, 8).value,
            'gender': sheet.cell(row, 9).value,
            'unit': sheet.cell(row, 11).value,
            'managerName': sheet.cell(row, 14).value,
            'managerEmail': sheet.cell(row, 15).value,
            'position': sheet.cell(row, 16).value,
            'email': sheet.cell(row, 18).value
        }
        cells.append(cell)

    with open('staff.json', 'w') as outfile:
        json.dump(cells, outfile)
    # print(cells)

    # save
    # wb.save(filename)


# print(cells)

process_workbook('staff.xlsx')
