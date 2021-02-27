import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import json


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['staff']

    cells = []

    for row in range(2, sheet.max_row + 1):
        cell = {

            'name': sheet.cell(row, 2).value,
            'email': sheet.cell(row, 3).value,

        }
        cells.append(cell)

    with open('people.js', 'w') as outfile:
        json.dump(cells, outfile)
    # print(cells)

    # save
    # wb.save(filename)


# print(cells)

process_workbook('staff.xlsx')
